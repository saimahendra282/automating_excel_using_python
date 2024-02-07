from openpyxl import load_workbook
import json
from datetime import datetime
wb = load_workbook('sales_data.xlsx')
ws = wb.active
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Convert datetime objects to strings
    # here add the heading in the excel file
    row_data = {
        "OrderDate": row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else row[0],
        "Region": row[1],
        "Rep": row[2],
        "Item": row[3],
        "Units": row[4],
        "UnitCost": row[5],
        "Total": row[6],
        "Shipped": row[7]
    }
    data.append(row_data)
# Print the data in JSON format
print(json.dumps(data, indent=4))
