from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

data = {
    "Sai": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Mahi": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Sunny": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 92
    },
    "Teja": {
        "math": 30,
        "science": 25,
        "english": 45,
        "gym": 100
    },
    "John": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Sai'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# Applying borders to cells
border_style = Side(style='thin', color='000000')  # Thin border with black color
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

# Applying font to header cells
for col in range(1, len(headings) + 1):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="000000")

# Deleting row 7
ws.delete_rows(7)

# Save the workbook
wb.save("Sample.xlsx")
