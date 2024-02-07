from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the workbook
wb = load_workbook('sales_data.xlsx')
ws = wb.active

# Access cell A1 and print its value
cell_value = ws['A1'].value
print("Value of cell A1:", cell_value)
# Printing the total number of sheets in the workbook
print("Total sheets present in this file:", len(wb.sheetnames))
# Printing the headings of each column
print("Headings of each column:")
for idx, cell in enumerate(ws[1], 1):
    print(f"Column {idx}: {cell.value}")
# Adding the values in a specific column (column G)
total_cost = 0
for cell in ws['G']:
    if cell.value:
        try:
            total_cost += float(cell.value)
        except ValueError:
            pass  # Ignore non-numeric values
print("Total everyone's total cost:", total_cost)
# Finding the average of the values in a specific column (column G)
column_to_average = 'G'
column_values = []
for cell in ws[column_to_average]:
    if cell.value and isinstance(cell.value, (int, float)):
        column_values.append(cell.value)
if column_values:
    average = sum(column_values) / len(column_values)
    print(f"Average cost of everyone in column {column_to_average}: {average}")
else:
    print(f"No numeric values found in column {column_to_average}")
# Renaming the cells
cell = ws['A14']
previous_value = cell.value
cell.value = "Testing Value"  # Pass the empty value here for removing the name
print(f"Cell A14 renamed from '{previous_value}' to '{cell.value}'")
# Adding rows
new_row_data = ['new', 'row', 'is', 'added', '!']
ws.append(new_row_data)
print("New row added:", new_row_data)
# Adding a column
new_column_data = ['new', 'column', 'is', 'added', '!']
new_column_letter = get_column_letter(ws.max_column + 1)
for idx, value in enumerate(new_column_data, start=2):
    ws[new_column_letter + str(idx)] = value 
print("New column added:", new_column_data)

# Deleting rows
row_to_delete = 8  # Example: Delete row 8
ws.delete_rows(row_to_delete)
print(f"Row {row_to_delete} deleted")

# Deleting a column
column_to_delete = 'J'  # Example: Delete column J
ws.delete_cols(ws[1].index(column_to_delete))
print(f"Column {column_to_delete} deleted")
# for merging 
start_row = 1
start_column = 1
end_row = 1
end_column = 5
ws.merge_cells(start_row=start_row, start_column=start_column,
               end_row=end_row, end_column=end_column)
print("Merge operation done")
# undoing merge
start_row = 1
start_column = 1
end_row = 1
end_column = 5
ws.unmerge_cells(start_row=start_row, start_column=start_column,
               end_row=end_row, end_column=end_column)

wb.save('sales_data.xlsx')
print("Data modified successfully")
